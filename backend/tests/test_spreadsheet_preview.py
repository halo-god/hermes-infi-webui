"""Spreadsheet preview styling: xlsx/csv extractors must emit a complete
styled HTML document (the sandboxed iframe cannot see app CSS)."""
from __future__ import annotations

import io

from app.core.files import extract_csv_html, extract_xlsx_html


def _workbook_bytes() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "销售数据"
    ws.append(["产品", "数量", "单价"])
    ws.append([" 显示屏", 120, 1999.5])
    ws.append(["控制器", 45, 350])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_html_is_a_styled_document():
    html = extract_xlsx_html(_workbook_bytes())
    assert html is not None
    assert html.startswith("<!DOCTYPE html>")
    assert "<style>" in html
    # Spreadsheet look: header row + zebra rows + right-aligned numbers.
    assert "<thead>" in html and "<th>产品</th>" in html
    assert 'class="num"' in html
    assert "销售数据" in html  # sheet name kept as a section heading


def test_xlsx_html_fake_workbook_error_is_styled():
    html = extract_xlsx_html(b"<html>not a real xlsx</html>")
    assert html is not None
    assert html.startswith("<!DOCTYPE html>")
    assert "不是有效的 .xlsx" in html


def test_csv_html_is_a_styled_document():
    html = extract_csv_html("名称,数量\nA,1\nB,2\n".encode())
    assert html is not None
    assert html.startswith("<!DOCTYPE html>")
    assert "<style>" in html
    assert "<th>名称</th>" in html
    assert "<td>A</td>" in html


def test_csv_html_empty_returns_none():
    assert extract_csv_html(b"") is None


def _styled_workbook_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "报表"
    ws["A1"] = "区域"
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="B8852A")
    ws["B1"] = "完成率"
    ws["B1"].font = Font(bold=True)
    ws["A2"] = "华东"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["B2"] = 0.847
    ws["B2"].number_format = "0.0%"
    ws["A3"] = "华南"
    ws["B3"] = 1234567.891
    ws["B3"].number_format = "#,##0.00"
    ws["A4"] = "合计（合并说明行）"
    ws.merge_cells("A4:B4")
    ws.column_dimensions["A"].width = 22
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "B2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_rich_render_restores_styles():
    html = extract_xlsx_html(_styled_workbook_bytes())
    assert html is not None
    # Column widths via colgroup, row heights via tr style.
    assert "<colgroup>" in html and "width:" in html
    assert 'style="height:24px"' in html
    # Fill + font color + bold from the header cell.
    assert "background:#b8852a" in html.lower()
    assert "color:#ffffff" in html.lower()
    assert "font-weight:700" in html
    # Alignment from A2.
    assert "text-align:center" in html
    # Merged range renders as rowspan/colspan on the anchor cell.
    assert 'rowspan="1" colspan="2"' in html
    # Number formats applied.
    assert "84.7%" in html
    assert "1,234,567.89" in html
    # Frozen first column gets the sticky class.
    assert 'class="fc"' in html or "fc" in html


def test_is_legacy_spreadsheet_html():
    from app.core.files import is_legacy_spreadsheet_html
    assert is_legacy_spreadsheet_html("xlsx", "<h3>Sheet</h3><table><tr>") is True
    assert is_legacy_spreadsheet_html("csv", "<table><tr><th>a</th>") is True
    assert is_legacy_spreadsheet_html("xlsx", "<!DOCTYPE html><html>") is False
    assert is_legacy_spreadsheet_html("md", "<table>") is False
    assert is_legacy_spreadsheet_html("xlsx", None) is False
